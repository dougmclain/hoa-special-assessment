from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.http import HttpResponse
from .models import Association, SpecialAssessment, Unit, UnitAssessment, AdditionalFee, Payment
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from decimal import Decimal


class AdditionalFeeInline(admin.TabularInline):
    model = AdditionalFee
    extra = 0
    readonly_fields = ('monthly_payment',)
    fields = ('fee_type', 'fee_amount', 'monthly_payment', 'description')


class PaymentInline(admin.TabularInline):
    model = Payment
    extra = 0
    fields = ('payment_date', 'amount', 'payment_method', 'reference_number', 'notes')
    ordering = ['-payment_date']


@admin.register(Association)
class AssociationAdmin(admin.ModelAdmin):
    list_display = ('name', 'management_company', 'unit_count', 'created_at')
    search_fields = ('name', 'management_company')
    readonly_fields = ('created_at', 'updated_at')

    def unit_count(self, obj):
        return obj.units.count()
    unit_count.short_description = 'Number of Units'


@admin.register(SpecialAssessment)
class SpecialAssessmentAdmin(admin.ModelAdmin):
    list_display = ('name', 'association', 'total_loan_amount', 'interest_rate', 'start_date', 'loan_period_months')
    list_filter = ('association', 'start_date')
    search_fields = ('name', 'association__name')
    readonly_fields = ('created_at', 'updated_at')
    fieldsets = (
        ('Basic Information', {
            'fields': ('association', 'name', 'description', 'start_date')
        }),
        ('Loan Information', {
            'fields': ('total_loan_amount', 'interest_rate', 'loan_period_months', 'monthly_loan_payment')
        }),
        ('Assessment Totals', {
            'fields': ('total_base_assessment', 'total_lce_assessments')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    actions = ['export_to_excel']

    def export_to_excel(self, request, queryset):
        """Export selected special assessment(s) to Excel"""
        if queryset.count() != 1:
            self.message_user(request, "Please select exactly one special assessment to export.", level='error')
            return

        assessment = queryset.first()

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Assessment Summary"

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Assessment Information
        ws['A1'] = assessment.association.name
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = assessment.name
        ws['A2'].font = Font(bold=True, size=12)

        ws['A4'] = "Loan Amount:"
        ws['B4'] = float(assessment.total_loan_amount)
        ws['B4'].number_format = '"$"#,##0.00'

        ws['A5'] = "Interest Rate:"
        ws['B5'] = float(assessment.interest_rate) / 100
        ws['B5'].number_format = '0.00%'

        ws['A6'] = "Loan Period:"
        ws['B6'] = f"{assessment.loan_period_months} months"

        ws['A7'] = "Monthly Loan Payment:"
        ws['B7'] = float(assessment.monthly_loan_payment)
        ws['B7'].number_format = '"$"#,##0.00'

        # Unit Assessment Table
        row = 10
        headers = ['Unit', 'Base Assessment', 'LCE Fees', 'Total Assessment', 'Monthly Base', 'Monthly LCE', 'Total Monthly', 'Total Paid', 'Balance', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        row += 1
        for unit_assessment in assessment.unit_assessments.all():
            ws.cell(row=row, column=1, value=unit_assessment.unit.unit_number)
            ws.cell(row=row, column=2, value=float(unit_assessment.base_assessment_amount)).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=3, value=float(unit_assessment.total_lce_fees())).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=4, value=float(unit_assessment.total_assessment_amount())).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=5, value=float(unit_assessment.monthly_base_payment)).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=6, value=float(unit_assessment.total_lce_monthly_payment())).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=7, value=float(unit_assessment.total_monthly_payment())).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=8, value=float(unit_assessment.total_paid())).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=9, value=float(unit_assessment.remaining_balance())).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=10, value=unit_assessment.payment_status())
            row += 1

        # Adjust column widths
        for col in range(1, 11):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{assessment.name.replace(" ", "_")}.xlsx"'
        wb.save(response)
        return response

    export_to_excel.short_description = "Export to Excel"


@admin.register(Unit)
class UnitAdmin(admin.ModelAdmin):
    list_display = ('unit_number', 'association', 'owner_name', 'common_expense_allocation')
    list_filter = ('association',)
    search_fields = ('unit_number', 'owner_name', 'owner_email')
    readonly_fields = ('created_at', 'updated_at')


@admin.register(UnitAssessment)
class UnitAssessmentAdmin(admin.ModelAdmin):
    list_display = ('unit_number', 'special_assessment', 'total_assessment_display', 'total_monthly_display', 'total_paid_display', 'balance_display', 'status_display')
    list_filter = ('special_assessment', 'payment_option', 'unit__association')
    search_fields = ('unit__unit_number', 'unit__owner_name')
    readonly_fields = ('monthly_base_payment', 'total_assessment_display', 'total_monthly_display', 'total_paid_display', 'balance_display', 'created_at', 'updated_at')
    inlines = [AdditionalFeeInline, PaymentInline]

    fieldsets = (
        ('Unit and Assessment', {
            'fields': ('unit', 'special_assessment')
        }),
        ('Assessment Details', {
            'fields': ('base_assessment_amount', 'payment_option', 'monthly_base_payment')
        }),
        ('Summary', {
            'fields': ('total_assessment_display', 'total_monthly_display', 'total_paid_display', 'balance_display'),
            'classes': ('wide',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def unit_number(self, obj):
        return obj.unit.unit_number
    unit_number.short_description = 'Unit'
    unit_number.admin_order_field = 'unit__unit_number'

    def total_assessment_display(self, obj):
        return format_html('<strong>${:,.2f}</strong>', obj.total_assessment_amount())
    total_assessment_display.short_description = 'Total Assessment'

    def total_monthly_display(self, obj):
        return format_html('${:,.2f}', obj.total_monthly_payment())
    total_monthly_display.short_description = 'Total Monthly Payment'

    def total_paid_display(self, obj):
        return format_html('<span style="color: green;">${:,.2f}</span>', obj.total_paid())
    total_paid_display.short_description = 'Total Paid'

    def balance_display(self, obj):
        balance = obj.remaining_balance()
        color = 'red' if balance > 0 else 'green'
        return format_html('<span style="color: {};">${:,.2f}</span>', color, balance)
    balance_display.short_description = 'Balance'

    def status_display(self, obj):
        status = obj.payment_status()
        colors = {
            'Paid in Full': 'green',
            'Current': 'blue',
            'Behind': 'orange',
            'Not Started': 'red',
            'Partial Payment': 'orange',
            'Not Paid': 'red',
        }
        color = colors.get(status, 'black')
        return format_html('<span style="color: {}; font-weight: bold;">{}</span>', color, status)
    status_display.short_description = 'Status'


@admin.register(AdditionalFee)
class AdditionalFeeAdmin(admin.ModelAdmin):
    list_display = ('unit_number', 'fee_type', 'fee_amount', 'monthly_payment')
    list_filter = ('fee_type', 'unit_assessment__special_assessment')
    search_fields = ('unit_assessment__unit__unit_number', 'fee_type')
    readonly_fields = ('monthly_payment', 'created_at', 'updated_at')

    def unit_number(self, obj):
        return obj.unit_assessment.unit.unit_number
    unit_number.short_description = 'Unit'


@admin.register(Payment)
class PaymentAdmin(admin.ModelAdmin):
    list_display = ('unit_number', 'payment_date', 'amount', 'payment_method', 'reference_number')
    list_filter = ('payment_date', 'payment_method', 'unit_assessment__special_assessment')
    search_fields = ('unit_assessment__unit__unit_number', 'reference_number')
    date_hierarchy = 'payment_date'
    readonly_fields = ('created_at', 'updated_at')

    def unit_number(self, obj):
        return obj.unit_assessment.unit.unit_number
    unit_number.short_description = 'Unit'
    unit_number.admin_order_field = 'unit_assessment__unit__unit_number'
